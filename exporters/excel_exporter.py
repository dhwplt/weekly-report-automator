import os
import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import config
from core.models import DailyReportBatch, WeeklyReportItem

def get_week_range_str() -> str:
    today = datetime.date.today()
    start = today - datetime.timedelta(days=today.weekday())
    end = start + datetime.timedelta(days=4)
    week_num = today.isocalendar()[1]
    return f"Week {week_num} ({start.strftime('%d-%m')} to {end.strftime('%d-%m')})"

def apply_header_styles(ws, week_str: str):
    ws.merge_cells('A1:H1')
    banner = ws['A1']
    banner.value = f"Weekly Report ( {week_str.split('(')[1].split(')')[0]} )"
    banner.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    banner.font = Font(bold=True, color="000000")
    banner.alignment = Alignment(horizontal='center', vertical='center')
    
    thick_border = Border(
        left=Side(style='thick'), right=Side(style='thick'),
        top=Side(style='thick'), bottom=Side(style='thick')
    )
    for row in ws['A1:H1']:
        for cell in row:
            cell.border = thick_border
            
    headers = [
        "Date", "Website/System", "Task / Scope", "Priority", 
        "Deadline", "Status", "Progress (%)", "Remarks"
    ]
    
    header_fill = PatternFill(start_color="8EA9DB", end_color="8EA9DB", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

def apply_row_styles(ws, row_idx: int, item: WeeklyReportItem):
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    ws.cell(row=row_idx, column=1).value = item.date
    ws.cell(row=row_idx, column=2).value = item.website_or_system
    ws.cell(row=row_idx, column=3).value = item.task_scope
    ws.cell(row=row_idx, column=4).value = item.priority
    ws.cell(row=row_idx, column=5).value = item.deadline
    ws.cell(row=row_idx, column=6).value = item.status
    ws.cell(row=row_idx, column=7).value = item.progress_percentage
    ws.cell(row=row_idx, column=8).value = item.remarks

    status_lower = item.status.lower()
    if "completed" in status_lower:
        fill_color = "E2EFDA"
        font_color = "006100"
    elif "in progress" in status_lower:
        fill_color = "FFF2CC"
        font_color = "9C5700"
    elif "on hold" in status_lower or "blocked" in status_lower:
        fill_color = "FCE4D6"
        font_color = "9C0006"
    else:
        fill_color = "FFFFFF"
        font_color = "000000"

    status_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    status_font = Font(color=font_color)

    for col_idx in range(1, 9):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border = thin_border
        
        if col_idx in [3, 8]:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        else:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
        if col_idx == 6:
            cell.fill = status_fill
            cell.font = status_font

def auto_adjust_columns(ws):
    min_width = 10
    max_width = 50
    
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        column = get_column_letter(col_idx)
        
        for cell in ws[column]:
            try:
                if cell.value:
                    lines = str(cell.value).split('\n')
                    longest_line = max([len(l) for l in lines]) if lines else 0
                    if longest_line > max_length:
                        max_length = longest_line
            except:
                pass
                
        adjusted_width = (max_length + 2)
        if adjusted_width < min_width:
            adjusted_width = min_width
        elif adjusted_width > max_width:
            adjusted_width = max_width
            
        ws.column_dimensions[column].width = adjusted_width

def export_to_excel(batch: DailyReportBatch):
    if not batch.items:
        return
        
    excel_path = config.EXCEL_OUTPUT_PATH
    week_str = get_week_range_str()
    
    if excel_path.exists():
        wb = load_workbook(excel_path)
    else:
        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
            
    if week_str not in wb.sheetnames:
        ws = wb.create_sheet(title=week_str)
        apply_header_styles(ws, week_str)
    else:
        ws = wb[week_str]
        
    existing_entries = {}
    for row_idx in range(3, ws.max_row + 1):
        date_val = ws.cell(row=row_idx, column=1).value
        sys_val = ws.cell(row=row_idx, column=2).value
        if date_val and sys_val:
            existing_entries[(str(date_val), str(sys_val))] = row_idx
            
    next_new_row = ws.max_row + 1
    if ws.max_row == 2 and not ws.cell(row=3, column=1).value:
        next_new_row = 3
        
    for item in batch.items:
        key = (item.date, item.website_or_system)
        if key in existing_entries:
            row_idx = existing_entries[key]
        else:
            row_idx = next_new_row
            next_new_row += 1
            
        apply_row_styles(ws, row_idx, item)
        
    auto_adjust_columns(ws)
    
    wb.save(excel_path)
    print(f"Exported {len(batch.items)} items to {excel_path}")
